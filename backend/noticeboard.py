"""Fast and defensive DIU Notice Board discovery.

Design goals:
- Mid/Final only; Spring/Summer/Fall; any year.
- Use the rendered DIU notice cards instead of guessing from raw HTML.
- Rank likely CSE routine/seat-plan notices before opening detail pages.
- Use ONE Playwright browser/context per request.
- Download/parse only a small number of likely attachments.
- Never parse the whole seat-plan PDF during discovery more than once.
- A missing seat plan is a valid result, not an error.
"""
from __future__ import annotations

import io
import json
import logging
import os
import re
import tempfile
import time
from typing import Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# Silence pdfminer/pdfplumber's noisy CropBox warnings. They are harmless for
# the DIU PDFs and should never flood Render/local logs.
for _name in ("pdfminer", "pdfminer.pdfpage", "pdfplumber"):
    logging.getLogger(_name).setLevel(logging.ERROR)

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
except Exception:  # pragma: no cover
    sync_playwright = None
    PlaywrightTimeout = Exception

from parser import (
    build_student_routine,
    clean_course_code,
    normalize_batch,
    normalize_section,
    parse_exam_routine,
    parse_seat_plan,
)

NOTICEBOARD_URL = "https://daffodilvarsity.edu.bd/noticeboard"
MAX_DOCUMENT_BYTES = 30 * 1024 * 1024
HTTP_TIMEOUT = (8, 20)
PAGE_TIMEOUT = 22_000
CANDIDATE_CACHE_SECONDS = 600
SEMESTERS = ("spring", "summer", "fall")
EXAM_TYPES = ("mid", "final")

# Shared caches: avoid re-crawling the notice board for every student click.
_CANDIDATE_CACHE: dict[tuple, tuple[float, list]] = {}

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/151 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/json,application/pdf," \
              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
}


def norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "")).strip()


def absolute(url: str, base: str = NOTICEBOARD_URL) -> str:
    u = str(url or "").strip()
    # JSON/API payloads sometimes emit escaped slashes like https:\/\/host\/path
    u = u.replace("\\/", "/").replace("\\\\/", "/")
    # Drop accidental double-scheme artifacts such as https://host/https://host/...
    u = re.sub(r"^(https?://[^/]+)/(https?://)", r"\2", u, flags=re.I)
    return urljoin(base, u)


def is_document(url: str) -> bool:
    p = urlparse(url).path.lower()
    q = url.lower()
    return p.endswith((".pdf", ".xlsx", ".xlsm", ".xltx", ".xltm")) or any(
        x in q for x in (".pdf?", ".xlsx?", ".xlsm?", ".xltx?", ".xltm?")
    )


def is_spreadsheet(url: str) -> bool:
    p = urlparse(url).path.lower()
    q = url.lower()
    return p.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")) or any(
        x in q for x in (".xlsx?", ".xlsm?", ".xltx?", ".xltm?")
    )


def normalize_semester(v: str) -> str:
    s = norm(v).lower()
    return next((x for x in SEMESTERS if re.search(rf"\b{x}\b", s)), "")


def normalize_exam_type(v: str) -> str:
    s = norm(v).lower().replace("–", "-").replace("—", "-")
    if re.search(r"\bmid(?:term)?\b|mid[- ]?semester", s):
        return "mid"
    if re.search(r"\bfinal\b|final[- ]?(?:exam|examination|semester)", s):
        return "final"
    return ""


def session_matches(text: str, semester: str, year: Optional[int]) -> bool:
    s = norm(text).lower().replace("-", " ").replace("_", " ")
    if not re.search(rf"\b{re.escape(semester)}\b", s):
        return False
    if year is None:
        return True
    y = int(year)
    return bool(re.search(rf"\b{y}\b", s) or re.search(rf"\b{y % 100:02d}\b", s))


def is_cse(text: str) -> bool:
    s = norm(text).lower()
    return (
        "computer science and engineering" in s
        or "computer science & engineering" in s
        or "department of computer science" in s
        or bool(re.search(r"\bcse\b", s))
        or bool(re.search(r"\bcse\d{3,4}\b", s))
    )


def kind_of_notice(text: str) -> str:
    s = norm(text).lower()
    if re.search(r"seat\s*(?:plan|details)|sitting\s*plan", s):
        return "seat"
    if re.search(r"\broutine\b|\bschedule\b", s):
        return "routine"
    return "other"


def _candidate_score(item: dict, exam_type: str, semester: str, year: Optional[int], kind: str, section: str = "") -> int:
    text = norm(" ".join(str(item.get(k, "")) for k in (
        "title", "context", "detail_title", "detail_text", "url"
    ))).lower()
    score = 0

    # Cheap metadata filter first. This is what makes the crawler fast.
    if normalize_exam_type(text) == exam_type:
        score += 1000
    elif exam_type == "final" and "final" in text:
        score += 450
    elif exam_type == "mid" and re.search(r"\bmid", text):
        score += 450
    else:
        return -100000

    if session_matches(text, semester, year):
        score += 650
    elif year and str(year) in text:
        score += 250
    else:
        return -100000

    k = kind_of_notice(text)
    if kind == "routine":
        if k == "routine": score += 600
        if "updated" in text: score += 400
        if "seat" in text: score -= 800
    elif kind == "seat":
        if k == "seat": score += 900
        if "seat plan" in text: score += 250
        if "seat details" in text: score += 150
        if k == "routine": score -= 800

    if is_cse(text): score += 550
    if section and normalize_section(section) in normalize_section(text): score += 250
    if "examination" in text or "exam" in text: score += 100
    # Prefer real notice / noticeFile URLs over API noise or broken escapes.
    u = norm(item.get("url", "")).lower()
    if "noticefile" in u or u.endswith((".pdf", ".xlsx")):
        score += 200
    if "\\/" in u or "api/v1/public" in u:
        score -= 1500
    if kind == "routine" and "updated" in text and is_cse(text):
        score += 350
    return score


def _extract_urls(value: str, base: str) -> list[str]:
    if not value:
        return []
    patterns = [
        r"https?://[^\s\"'<>]+",
        r"/noticeboard/[^\s\"'<>]+",
        r"/notice_detail/[^\s\"'<>]+",
        r"/download-file/[^\s\"'<>]+",
        r"/noticeFile/[^\s\"'<>]+",
        r"[^\s\"'<>]+\.(?:pdf|xlsx|xlsm|xltx|xltm)(?:\?[^\s\"'<>]*)?",
    ]
    found = []
    for pattern in patterns:
        for raw in re.findall(pattern, str(value), re.I):
            u = absolute(raw.rstrip(".,;)]}>\\'\"`"), base)
            if u.startswith(("http://", "https://")):
                found.append(u)
    return list(dict.fromkeys(found))


def _notice_from_link(a, base=NOTICEBOARD_URL) -> Optional[dict]:
    href = a.get("href") or a.get("data-href") or a.get("data-url") or ""
    href = absolute(href, base)
    title = norm(a.get_text(" ", strip=True))
    if not href or len(title) < 4:
        return None
    if any(x in href.lower() for x in ("facebook.com", "instagram.com", "youtube.com", "mailto:")):
        return None
    parent = a.parent
    context = title
    # Current DIU cards contain title + department + category + date in the
    # surrounding card. Walk upward, but stop before the whole page.
    for _ in range(5):
        if parent is None:
            break
        t = norm(parent.get_text(" ", strip=True))
        if len(t) > len(context) and len(t) <= 1600:
            context = t
        parent = parent.parent
    return {"url": href, "title": title, "context": context, "source": NOTICEBOARD_URL}


def _dedupe(items: list[dict]) -> list[dict]:
    out = {}
    for x in items:
        u = x.get("url", "")
        if u:
            out[u] = x
    return list(out.values())


def _collect_dom(page) -> list[dict]:
    rows = page.locator("a[href], [data-href], [data-url]").evaluate_all(
        """
        els => els.map(a => {
          let p = a, context = (a.innerText || a.textContent || '').replace(/\\s+/g,' ').trim();
          for (let i=0; i<5 && p; i++, p=p.parentElement) {
            const t=(p.innerText||p.textContent||'').replace(/\\s+/g,' ').trim();
            if (t.length > context.length && t.length <= 1600) context=t;
          }
          return {
            href:a.getAttribute('href') || a.getAttribute('data-href') || a.getAttribute('data-url') || '',
            text:(a.innerText||a.textContent||'').replace(/\\s+/g,' ').trim(), context
          };
        })
        """
    )
    out=[]
    for r in rows:
        u=absolute(r.get("href",""))
        t=norm(r.get("text","")); c=norm(r.get("context",""))
        if not u or len(t)<4 or u.rstrip("/")==NOTICEBOARD_URL.rstrip("/"):
            continue
        if any(x in u.lower() for x in ("facebook.com","instagram.com","youtube.com","mailto:")):
            continue
        out.append({"url":u,"title":t,"context":c,"source":NOTICEBOARD_URL})
    return _dedupe(out)


def _collect_network_json(records: list[dict]) -> list[dict]:
    """Extract notice IDs/URLs from API responses without downloading files."""
    out=[]
    def walk(obj, inherited="", base=NOTICEBOARD_URL):
        if isinstance(obj, dict):
            scalars=[str(v) for v in obj.values() if isinstance(v,(str,int,float))]
            label=norm(" ".join(scalars))
            urls=[]
            for v in obj.values():
                if isinstance(v,str): urls.extend(_extract_urls(v,base))
            ident=None
            for k in ("id","notice_id","noticeId","noticeID","noticeNo","notice_no"):
                if k in obj and str(obj[k]).isdigit():
                    ident=str(obj[k]); break
            if ident:
                urls.append(absolute(f"/noticeboard/notice_detail/{ident}",base))
            for u in urls:
                out.append({"url":u,"title":label[:500] or "DIU notice","context":label[:1600],"source":base})
            for v in obj.values():
                walk(v,label,base)
        elif isinstance(obj,list):
            for v in obj: walk(v,inherited,base)

    for r in records:
        body=r.get("body","")
        ctype=r.get("content_type","").lower()
        if not body: continue
        if "json" in ctype or body.lstrip().startswith(("{","[")):
            try: walk(json.loads(body),base=r.get("url",NOTICEBOARD_URL))
            except Exception: pass
        for u in _extract_urls(body,r.get("url",NOTICEBOARD_URL)):
            out.append({"url":u,"title":"DIU notice","context":body[:1600],"source":r.get("url",NOTICEBOARD_URL)})
    return _dedupe(out)


def collect_notice_candidates(semester: str, year: Optional[int], exam_type: str) -> list[dict]:
    """One browser pass. Returns cheap notice metadata, not documents.

    Results are cached briefly so concurrent students share one crawl.
    """
    cache_key = (semester, year, exam_type)
    cached = _CANDIDATE_CACHE.get(cache_key)
    if cached:
        created, items = cached
        if time.monotonic() - created < CANDIDATE_CACHE_SECONDS and items:
            return [dict(x) for x in items]

    if sync_playwright is None:
        return []
    network = []
    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu", "--disable-extensions"],
        )
        context = browser.new_context(
            user_agent=HEADERS["User-Agent"],
            viewport={"width": 1280, "height": 900},
            accept_downloads=True,
        )
        page = context.new_page()
        page.set_default_timeout(12_000)

        def on_response(resp):
            try:
                c = (resp.headers.get("content-type") or "").lower()
                if any(x in c for x in ("json", "html", "text")):
                    cl = int(resp.headers.get("content-length") or 0)
                    if cl and cl > 2_000_000:
                        return
                    try:
                        network.append({
                            "url": resp.url,
                            "content_type": c,
                            "body": resp.text()[:2_000_000],
                        })
                    except Exception:
                        pass
            except Exception:
                pass

        page.on("response", on_response)
        try:
            page.goto(NOTICEBOARD_URL, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
            # Short settle only — do not wait full networkidle.
            try:
                page.wait_for_timeout(1200)
            except Exception:
                pass

            for selector in ('button:has-text("FSIT")', '[role="button"]:has-text("FSIT")'):
                try:
                    loc = page.locator(selector).first
                    if loc.count() and loc.is_visible():
                        loc.click(timeout=1000)
                        try:
                            page.wait_for_timeout(800)
                        except Exception:
                            pass
                        break
                except Exception:
                    pass

            # Lazy loading: few scrolls, stop when stable.
            previous = 0
            stable = 0
            for _ in range(5):
                count = page.locator("a[href]").count()
                if count == previous:
                    stable += 1
                else:
                    stable = 0
                previous = count
                if stable >= 2:
                    break
                page.mouse.wheel(0, 2200)
                page.wait_for_timeout(180)

            found = _collect_dom(page)
            found.extend(_collect_network_json(network))
            html_source = page.content()
            for u in _extract_urls(html_source, page.url):
                found.append({
                    "url": u,
                    "title": "DIU notice",
                    "context": html_source[:1400],
                    "source": NOTICEBOARD_URL,
                })

            found = _dedupe(found)
            explicit_visible_seats = _collect_explicit_seat_candidates(page, exam_type, semester, year)
            found = _dedupe(found + explicit_visible_seats)

            ranked = []
            for n in found:
                score = max(
                    _candidate_score(n, exam_type, semester, year, "routine"),
                    _candidate_score(n, exam_type, semester, year, "seat"),
                )
                if score > -100000:
                    n["score_hint"] = score
                    ranked.append(n)
            ranked.sort(key=lambda x: x.get("score_hint", 0), reverse=True)

            explicit_seats = [
                n for n in ranked
                if kind_of_notice(norm(f"{n.get('title','')} {n.get('context','')}")) == "seat"
                and is_cse(norm(f"{n.get('title','')} {n.get('context','')}"))
            ][:12]
            explicit_seats += explicit_visible_seats
            pool = _dedupe(ranked[:40] + explicit_seats)
            pool.sort(key=lambda x: x.get("score_hint", 0), reverse=True)
            result = pool[:55]
            _CANDIDATE_CACHE[cache_key] = (time.monotonic(), [dict(x) for x in result])
            # Bound cache size
            if len(_CANDIDATE_CACHE) > 12:
                oldest = min(_CANDIDATE_CACHE.items(), key=lambda kv: kv[1][0])[0]
                _CANDIDATE_CACHE.pop(oldest, None)
            return result
        finally:
            browser.close()



def static_collect_notices(session: requests.Session, semester: str, year: Optional[int], exam_type: str) -> list[dict]:
    """Fallback for servers where Playwright cannot start."""
    try:
        r=session.get(NOTICEBOARD_URL,headers=HEADERS,timeout=HTTP_TIMEOUT)
        r.raise_for_status()
    except Exception:
        return []
    soup=BeautifulSoup(r.text,"html.parser")
    found=[]
    for a in soup.find_all("a",href=True):
        n=_notice_from_link(a)
        if not n: continue
        score=max(_candidate_score(n,exam_type,semester,year,"routine"),_candidate_score(n,exam_type,semester,year,"seat"))
        if score>-100000:
            n["score_hint"]=score; found.append(n)
    found.sort(key=lambda x:x.get("score_hint",0),reverse=True)
    for u in _seat_notice_fallback_urls(exam_type, semester, year):
        found.append({
            "url":u,
            "title":f"{exam_type.title()} Examination Seat Plan ({semester.title()}-{year})",
            "context":f"CSE {exam_type} examination seat plan {semester} {year}",
            "source":NOTICEBOARD_URL,
            "score_hint":100,
        })
    return _dedupe(found)[:45]


def _read_document(raw: bytes, final_url: str) -> tuple[str, str]:
    """Cheap text sniffing. Full parsing is done only after ranking."""
    if not raw or len(raw)>MAX_DOCUMENT_BYTES:
        raise ValueError("Document is missing or larger than 30 MB")
    low=final_url.lower()
    if raw.startswith(b"%PDF") or ".pdf" in low:
        # Use pypdf when available; it does not emit pdfminer CropBox warnings.
        try:
            from pypdf import PdfReader
            reader=PdfReader(io.BytesIO(raw), strict=False)
            parts=[]
            for p in reader.pages[:8]:
                try: parts.append(p.extract_text() or "")
                except Exception: pass
            return norm("\n".join(parts)), "pdf"
        except Exception:
            # Do not fail discovery merely because text extraction is imperfect.
            return "", "pdf"
    if raw[:2]==b"PK" or is_spreadsheet(final_url):
        wb=load_workbook(io.BytesIO(raw),data_only=True,read_only=True)
        parts=[]
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                parts.append(" ".join(norm(v) for v in row if v is not None))
        return norm("\n".join(parts)), "xlsx"
    raise ValueError("Unsupported document")


def download_document(session: requests.Session, url: str) -> Optional[dict]:
    try:
        r=session.get(url,headers=HEADERS,timeout=HTTP_TIMEOUT,allow_redirects=True)
        if not r.ok or len(r.content)>MAX_DOCUMENT_BYTES:
            return None
        text,ft=_read_document(r.content,r.url)
        return {"url":url,"final_url":r.url,"bytes":r.content,"file_type":ft,"text_content":text}
    except Exception:
        return None



def _download_with_playwright(context, url: str, referer: str = NOTICEBOARD_URL) -> Optional[dict]:
    """Download an attachment through the browser, with a navigation fallback."""
    headers={"Referer": referer, "Accept": "application/pdf,application/octet-stream,*/*"}

    # 1) Browser request context: fast and cookie-aware.
    try:
        resp = context.request.get(url, timeout=18_000, headers=headers)
        if resp.ok:
            raw = resp.body()
            if raw and len(raw) <= MAX_DOCUMENT_BYTES:
                final_url = resp.url or url
                text, ft = _read_document(raw, final_url)
                return {"url": url, "final_url": final_url, "bytes": raw,
                        "file_type": ft, "text_content": text}
    except Exception:
        pass

    # 2) Some DIU download endpoints only return the PDF when navigated like a
    # real browser page. Capture the main-document response body.
    page=None
    try:
        page=context.new_page()
        response=page.goto(url, wait_until="commit", timeout=18_000)
        if response is not None and response.ok:
            raw=response.body()
            final_url=response.url or url
            if raw and len(raw) <= MAX_DOCUMENT_BYTES:
                text, ft = _read_document(raw, final_url)
                return {"url": url, "final_url": final_url, "bytes": raw,
                        "file_type": ft, "text_content": text}
    except Exception:
        pass
    finally:
        try:
            if page: page.close()
        except Exception:
            pass
    return None


def _capture_view_download(page, context) -> list[dict]:
    """Capture documents exposed by DIU's View/Download UI.

    The notice details page has changed over time. The control may be a
    normal anchor, a JS button, a download event, or a popup. We handle all
    four without assuming a particular HTML implementation.
    """
    captured = []
    controls = page.locator("a,button").filter(
        has_text=re.compile(r"view\s*/?\s*download|view\s+download|download", re.I)
    )
    try:
        count = min(controls.count(), 4)
    except Exception:
        count = 0

    for i in range(count):
        control = controls.nth(i)
        try:
            if not control.is_visible():
                continue
        except Exception:
            continue

        before_url = page.url
        popup = None
        responses = []

        def on_response(resp):
            try:
                ct = (resp.headers.get("content-type") or "").lower()
                u = resp.url.lower()
                if ("pdf" in ct or "spreadsheet" in ct or is_document(u)
                        or "/download-file/" in u or "noticefile" in u):
                    responses.append(resp)
            except Exception:
                pass

        page.on("response", on_response)
        try:
            try:
                with page.expect_popup(timeout=2500) as pop_info:
                    control.click(timeout=2500)
                popup = pop_info.value
                try:
                    popup.wait_for_load_state("domcontentloaded", timeout=7000)
                except Exception:
                    pass
            except Exception:
                try:
                    with page.expect_download(timeout=6000) as dl_info:
                        control.click(timeout=2500)
                    dl = dl_info.value
                    path = dl.path()
                    if path:
                        raw = open(path, "rb").read()
                        filename = dl.suggested_filename or "attachment.pdf"
                        text, ft = _read_document(raw, filename)
                        captured.append({"url": before_url, "final_url": before_url,
                                         "bytes": raw, "file_type": ft,
                                         "text_content": text})
                        continue
                except Exception:
                    try:
                        control.click(timeout=2500)
                    except Exception:
                        continue

            if popup is not None:
                # The popup may itself be the PDF or may contain another
                # View/Download link.
                if is_document(popup.url) or "download-file/" in popup.url.lower() or "noticefile" in popup.url.lower():
                    d = _download_with_playwright(context, popup.url)
                    if d:
                        captured.append(d)
                else:
                    try:
                        for a in _extract_detail_attachments(popup):
                            d = _download_with_playwright(context, a["url"])
                            if d:
                                captured.append(d)
                    except Exception:
                        pass
                    try:
                        popup.close()
                    except Exception:
                        pass

            # Finally use document-like network responses produced by the click.
            for resp in responses:
                try:
                    raw = resp.body()
                    if not raw or len(raw) > MAX_DOCUMENT_BYTES:
                        continue
                    text, ft = _read_document(raw, resp.url)
                    captured.append({"url": resp.url, "final_url": resp.url,
                                     "bytes": raw, "file_type": ft,
                                     "text_content": text})
                except Exception:
                    continue
        finally:
            try:
                page.remove_listener("response", on_response)
            except Exception:
                pass

    # Deduplicate by final URL/size.
    unique = []
    seen = set()
    for d in captured:
        key = (d.get("final_url"), len(d.get("bytes", b"")), d.get("file_type"))
        if key in seen:
            continue
        seen.add(key)
        unique.append(d)
    return unique[:8]


def _extract_detail_attachments(page) -> list[dict]:
    rows=page.locator("a[href],iframe[src],embed[src],object[data],[data-href],[data-url],button,[onclick]").evaluate_all(
        """
        els => els.map(el => ({
          href:el.getAttribute('href')||el.getAttribute('src')||el.getAttribute('data')||el.getAttribute('data-href')||el.getAttribute('data-url')||'',
          onclick:el.getAttribute('onclick')||'',
          text:(el.innerText||el.textContent||'').replace(/\\s+/g,' ').trim(),
          parent:el.parentElement?(el.parentElement.innerText||'').replace(/\\s+/g,' ').trim():''
        }))
        """
    )
    html_source=page.content()
    out=[]
    for row in rows:
        label=norm(f"{row.get('text','')} {row.get('parent','')}")
        for raw in (row.get("href",""),row.get("onclick","")):
            for u in _extract_urls(raw,page.url):
                if is_document(u) or "noticefile" in u.lower() or "/download-file/" in u.lower():
                    out.append({"url":u,"label":label})
    for u in _extract_urls(html_source,page.url):
        if is_document(u) or "noticefile" in u.lower() or "/download-file/" in u.lower():
            out.append({"url":u,"label":"attachment"})
    # Direct page URL may itself be a document.
    if is_document(page.url) or "noticefile" in page.url.lower() or "/download-file/" in page.url.lower():
        out.append({"url":page.url,"label":"attachment"})
    seen=set(); unique=[]
    for x in out:
        if x["url"] not in seen:
            seen.add(x["url"]); unique.append(x)
    return unique[:8]


def _routine_is_match(path: str, section: str, semester: str, year: Optional[int]) -> Optional[dict]:
    try:
        parsed=parse_exam_routine(path)
    except Exception:
        return None
    batch=normalize_batch(section)
    exams=[e for e in parsed.get("exams",[]) if normalize_batch(e.get("batch"))==batch]
    if not exams: return None
    if parsed.get("semester") and normalize_semester(parsed["semester"]) != semester: return None
    if year and parsed.get("year") and int(parsed["year"]) != int(year): return None
    cse=sum(1 for e in exams if str(e.get("course_code","")).upper().startswith("CSE"))
    if cse == 0: return None
    return parsed


def _save_temp(raw: bytes, ft: str) -> str:
    fd,path=tempfile.mkstemp(suffix=".xlsx" if ft=="xlsx" else ".pdf")
    os.close(fd)
    with open(path,"wb") as f: f.write(raw)
    return path


def _try_parse_routine_doc(d: dict, notice: dict, section: str, semester: str, year: Optional[int], score: int, detail_text: str = "") -> Optional[dict]:
    """Save a downloaded document, verify it is the requested CSE routine, return result or None."""
    if not d or not d.get("bytes"):
        return None
    path = _save_temp(d["bytes"], d["file_type"])
    try:
        parsed = _routine_is_match(path, section, semester, year)
    finally:
        try:
            os.remove(path)
        except OSError:
            pass
    if not parsed:
        return None
    return {
        **notice,
        **d,
        "detail_title": notice.get("title", ""),
        "detail_text": detail_text,
        "parsed_routine": parsed,
        "score": score,
    }


def _inspect_routine_documents(session, candidates, section, exam_type, semester, year):
    """Open only top routine notices and parse only their attachments.

    Handles three real DIU cases:
    1. Candidate URL is already a PDF/XLSX (noticeFile / download-file).
    2. Candidate is a notice detail page with a View/Download control.
    3. Candidate is a detail page whose HTML embeds a noticeFile link.

    Strategy: try direct document URLs first (fast, no page render), then
    open only a few high-scoring CSE notice detail pages.
    """
    def score_of(n):
        return _candidate_score(n, exam_type, semester, year, "routine", section)

    ranked = sorted(candidates, key=score_of, reverse=True)
    direct = []
    pages = []
    for n in ranked:
        sc = score_of(n)
        if sc < 700:
            continue
        u = (n.get("url") or "").lower()
        text_meta = norm(f"{n.get('title','')} {n.get('context','')}")
        if is_document(u) or "noticefile" in u or "/download-file/" in u:
            bonus = 500 if ("cse" in u or is_cse(text_meta)) else 0
            direct.append((sc + bonus, n))
        else:
            pages.append((sc, n))
    direct.sort(key=lambda x: x[0], reverse=True)
    pages.sort(key=lambda x: x[0], reverse=True)

    best: list[dict] = []

    # Phase 1: direct document downloads (requests only, fast)
    for sc, notice in direct[:12]:
        d = download_document(session, notice["url"])
        hit = _try_parse_routine_doc(d, notice, section, semester, year, sc)
        if hit:
            best.append(hit)
            meta = norm(f"{notice.get('title','')} {notice.get('context','')} {notice.get('url','')}")
            if is_cse(meta) or "cse" in (notice.get("url") or "").lower():
                return best

    if sync_playwright is None:
        return best

    # Phase 2: detail pages via Playwright (only top CSE-ish)
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
        context = browser.new_context(user_agent=HEADERS["User-Agent"], accept_downloads=True)
        page = context.new_page()
        try:
            page_candidates = pages[:5]
            page_candidates.sort(
                key=lambda x: (
                    is_cse(norm(f"{x[1].get('title','')} {x[1].get('context','')}")),
                    x[0],
                ),
                reverse=True,
            )
            for sc, notice in page_candidates:
                url = notice.get("url", "")
                try:
                    page.goto(url, wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
                    try:
                        page.wait_for_timeout(400)
                    except Exception:
                        pass
                    try:
                        detail_text = norm(page.locator("body").inner_text())[:9000]
                    except Exception:
                        detail_text = ""

                    combined = norm(f"{notice.get('title','')} {detail_text}")
                    if kind_of_notice(combined) == "seat":
                        continue
                    if not is_cse(combined) and "cse" not in (url or "").lower():
                        if sc < 2500:
                            continue

                    attachments = _extract_detail_attachments(page)

                    def _att_key(a):
                        lab = (a.get("label") or "").lower()
                        uu = (a.get("url") or "").lower()
                        return (
                            ("cse" in lab or "cse" in uu),
                            ("download" in lab or "routine" in lab),
                            ("noticefile" in uu or is_document(uu)),
                        )

                    attachments = sorted(attachments, key=_att_key, reverse=True)[:6]

                    if not attachments:
                        try:
                            buttons = page.locator("button, a").filter(
                                has_text=re.compile(r"view\s*/?\s*download|download", re.I)
                            )
                            if buttons.count():
                                with page.expect_download(timeout=6000) as dl_info:
                                    buttons.first.click(timeout=2000)
                                dl = dl_info.value
                                path = dl.path()
                                if path:
                                    raw = open(path, "rb").read()
                                    final_url = getattr(dl, "suggested_filename", None) or page.url
                                    text_doc, ft = _read_document(raw, final_url)
                                    attachments.append({
                                        "url": page.url,
                                        "label": "view/download",
                                        "captured_bytes": raw,
                                        "captured_type": ft,
                                        "captured_text": text_doc,
                                    })
                        except Exception:
                            pass

                    for a in attachments:
                        if a.get("captured_bytes"):
                            d = {
                                "url": a.get("url", page.url),
                                "final_url": a.get("url", page.url),
                                "bytes": a["captured_bytes"],
                                "file_type": a["captured_type"],
                                "text_content": a.get("captured_text", ""),
                            }
                        else:
                            d = download_document(session, a["url"]) or _download_with_playwright(
                                context, a["url"], referer=page.url
                            )
                        hit = _try_parse_routine_doc(d, notice, section, semester, year, sc, detail_text)
                        if hit:
                            best.append(hit)
                            if is_cse(combined):
                                return best
                except Exception:
                    continue
        finally:
            browser.close()
    return best



def _slugify_notice_title(text: str) -> str:
    """Create a DIU-style notice slug from a visible notice title."""
    s = norm(text).lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def _seat_notice_fallback_urls(exam_type: str, semester: str, year: Optional[int]) -> list[str]:
    """Known/current DIU slug patterns used when the card is not exposed to JS.

    This is intentionally a small fallback list, not a broad crawler. The
    normal path is still the Notice Board DOM/API discovery.
    """
    if not year:
        return []
    y = int(year)
    e = "mid" if exam_type == "mid" else "final"
    patterns = [
        f"{e}-examination-seat-plan-{semester}-{y}-cse",
        f"{e}-exam-seat-plan-{semester}-{y}-cse",
        f"{e}-examination-seat-plan-{semester}-{y}",
        f"{e}-exam-seat-plan-{semester}-{y}",
        f"{e}-examination-seat-details-{semester}-{y}-cse",
        f"{e}-exam-seat-details-{semester}-{y}-cse",
    ]
    return [absolute(f"/noticeboard/{x}") for x in patterns]


def _collect_explicit_seat_candidates(page, exam_type: str, semester: str, year: Optional[int]) -> list[dict]:
    """Scan the visible notice board specifically for seat-plan cards.

    Do not reuse the general candidate score here: a seat-plan card can be
    ranked out by unrelated notices, pagination, or a slightly different
    title even though it is clearly the correct CSE notice.
    """
    out = []
    rows = page.locator("a[href], [data-href], [data-url]").evaluate_all(
        """
        els => els.map(a => {
          let p=a, parts=[];
          for(let i=0;i<6 && p;i++,p=p.parentElement){
            const t=(p.innerText||p.textContent||'').replace(/\\s+/g,' ').trim();
            if(t) parts.push(t);
          }
          return {
            href:a.getAttribute('href')||a.getAttribute('data-href')||a.getAttribute('data-url')||'',
            text:(a.innerText||a.textContent||'').replace(/\\s+/g,' ').trim(),
            context:parts.join(' | ')
          };
        })
        """
    )
    for r in rows:
        u = absolute(r.get("href", ""))
        title = norm(r.get("text", ""))
        context = norm(r.get("context", ""))
        text = norm(f"{title} {context} {u}")
        if not u or len(title) < 4:
            continue
        if not re.search(r"seat\s*(?:plan|details)|sitting\s*plan", text, re.I):
            continue
        if normalize_exam_type(text) != exam_type and not re.search(rf"\b{re.escape(exam_type)}\b", text, re.I):
            continue
        if year and not re.search(rf"\b{int(year)}\b|\b{int(year) % 100:02d}\b", text):
            continue
        if semester and not re.search(rf"\b{re.escape(semester)}\b", text, re.I):
            continue
        if not is_cse(text):
            continue
        out.append({"url":u,"title":title,"context":context,"source":NOTICEBOARD_URL,"score_hint":9999})
    return _dedupe(out)


def _inspect_notice_for_documents(page, context, notice: dict) -> list[dict]:
    """Open a notice and collect every plausible PDF/XLSX attachment."""
    docs = []
    try:
        page.goto(notice["url"], wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
    except Exception:
        return docs
    try:
        page.wait_for_timeout(700)
    except Exception:
        pass
    try:
        page.wait_for_load_state("networkidle", timeout=3500)
    except Exception:
        pass

    # Normal href/src/data/onclick attachments.
    for a in _extract_detail_attachments(page):
        d = _download_with_playwright(context, a["url"], referer=page.url)
        if d:
            docs.append(d)

    # JS View/Download control and popup/download/network response.
    docs.extend(_capture_view_download(page, context))

    # Last-resort: inspect inline scripts/HTML for escaped noticeFile URLs.
    try:
        html = page.content().replace("\\/", "/")
        for u in _extract_urls(html, page.url):
            if is_document(u) or "noticefile" in u.lower() or "/download-file/" in u.lower():
                d = _download_with_playwright(context, u, referer=page.url)
                if d:
                    docs.append(d)
    except Exception:
        pass

    seen=set(); unique=[]
    for d in docs:
        key=(d.get("final_url"),len(d.get("bytes",b"")),d.get("file_type"))
        if key not in seen:
            seen.add(key); unique.append(d)
    return unique

def _try_seat_doc(d, notice, section, exam_type, semester, year, batch, routine_keys, detail=""):
    """Parse a downloaded seat-plan PDF and verify section + routine overlap."""
    if not d or d.get("file_type") != "pdf" or not d.get("bytes"):
        return None
    path = _save_temp(d["bytes"], "pdf")
    try:
        try:
            alloc = parse_seat_plan(path)
        except Exception:
            return None
    finally:
        try:
            os.remove(path)
        except OSError:
            pass

    wanted = normalize_section(section)
    matches = [x for x in alloc if normalize_section(x.get("section")) == wanted and x.get("rooms")]
    if not matches:
        return None
    overlap = sum(
        1 for x in matches
        if (x.get("date"), x.get("slot"), clean_course_code(x.get("course_code", ""))) in routine_keys
    )
    if overlap <= 0:
        return None
    return {
        **notice,
        **d,
        "detail_title": notice.get("title", ""),
        "detail_text": detail,
        "seat_matches": matches,
        "seat_overlap": overlap,
        "score": _candidate_score(notice, exam_type, semester, year, "seat", section),
    }


def _inspect_seat_plan(session, candidates, routine, section, exam_type, semester, year):
    """Find the CSE seat-plan notice. Prefer direct PDF downloads (fast)."""
    batch = normalize_batch(section)
    routine_exams = routine.get("parsed_routine", {}).get("exams", [])
    routine_keys = {
        (e.get("date"), e.get("slot"), clean_course_code(e.get("course_code", "")))
        for e in routine_exams
        if normalize_batch(e.get("batch")) == batch
    }

    candidate_pool = []
    for n in candidates:
        text = norm(f"{n.get('title','')} {n.get('context','')} {n.get('url','')}")
        if re.search(r"seat\s*(?:plan|details)|sitting\s*plan", text, re.I):
            # Prefer CSE; still keep strong seat titles without CSE in title
            # (detail page / file name may carry it).
            candidate_pool.append(n)
    candidate_pool = _dedupe(candidate_pool)
    candidate_pool.sort(
        key=lambda n: _candidate_score(n, exam_type, semester, year, "seat", section),
        reverse=True,
    )

    seen_urls = {x.get("url") for x in candidate_pool}
    for u in _seat_notice_fallback_urls(exam_type, semester, year):
        if u not in seen_urls:
            candidate_pool.append({
                "url": u,
                "title": f"{exam_type.title()} Examination Seat Plan ({semester.title()}-{year})",
                "context": f"CSE {exam_type} examination seat plan {semester} {year}",
                "source": NOTICEBOARD_URL,
                "score_hint": 100,
            })
            seen_urls.add(u)

    # Phase 1: direct document URLs via requests only (no browser).
    direct = []
    pages = []
    for n in candidate_pool[:30]:
        u = (n.get("url") or "").lower()
        if is_document(u) or "noticefile" in u or "/download-file/" in u:
            direct.append(n)
        else:
            pages.append(n)

    # Prefer CSE-named seat PDFs first.
    direct.sort(
        key=lambda n: (
            "cse" in (n.get("url") or "").lower() or is_cse(norm(f"{n.get('title','')} {n.get('context','')}")),
            _candidate_score(n, exam_type, semester, year, "seat", section),
        ),
        reverse=True,
    )
    for notice in direct[:10]:
        d = download_document(session, notice["url"])
        hit = _try_seat_doc(d, notice, section, exam_type, semester, year, batch, routine_keys)
        if hit:
            return hit

    if sync_playwright is None:
        return None

    # Phase 2: only a few detail pages, short waits.
    pages.sort(
        key=lambda n: (
            is_cse(norm(f"{n.get('title','')} {n.get('context','')}")),
            _candidate_score(n, exam_type, semester, year, "seat", section),
        ),
        reverse=True,
    )
    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
        )
        context = browser.new_context(user_agent=HEADERS["User-Agent"], accept_downloads=True)
        page = context.new_page()
        page.set_default_timeout(10_000)
        try:
            for notice in pages[:6]:
                try:
                    page.goto(notice["url"], wait_until="domcontentloaded", timeout=PAGE_TIMEOUT)
                    try:
                        page.wait_for_timeout(500)
                    except Exception:
                        pass
                    try:
                        detail = norm(page.locator("body").inner_text())[:10000]
                    except Exception:
                        detail = ""

                    title_context = norm(" ".join([notice.get("title", ""), notice.get("context", ""), detail]))
                    if not re.search(r"seat\s*(?:plan|details)|sitting\s*plan", title_context, re.I):
                        continue
                    if not is_cse(title_context) and "cse" not in (notice.get("url") or "").lower():
                        continue
                    if year and not session_matches(title_context, semester, year):
                        # Allow if URL slug already matches session
                        if not session_matches(notice.get("url", ""), semester, year):
                            continue

                    # Prefer extracting attachment links over full click machinery.
                    attachments = _extract_detail_attachments(page)
                    for a in attachments[:5]:
                        d = download_document(session, a["url"])
                        if not d:
                            d = _download_with_playwright(context, a["url"], referer=page.url)
                        hit = _try_seat_doc(
                            d, notice, section, exam_type, semester, year, batch, routine_keys, detail
                        )
                        if hit:
                            return hit

                    # One View/Download attempt if needed.
                    docs = _capture_view_download(page, context)
                    for d in docs:
                        hit = _try_seat_doc(
                            d, notice, section, exam_type, semester, year, batch, routine_keys, detail
                        )
                        if hit:
                            return hit
                except Exception:
                    continue
        finally:
            browser.close()
    return None



def discover_documents(section: str, exam_type: str, semester: str, year: Optional[int]=None, include_seat_plan: bool=True) -> dict:
    section=normalize_section(section)
    exam_type=normalize_exam_type(exam_type)
    semester=normalize_semester(semester)
    if exam_type not in EXAM_TYPES: raise ValueError("Exam type must be Mid or Final.")
    if semester not in SEMESTERS: raise ValueError("Semester must be Spring, Summer, or Fall.")
    if year is not None and not 2000 <= int(year) <= 2100: raise ValueError("Invalid academic year.")

    session=requests.Session(); session.headers.update(HEADERS)
    candidates=collect_notice_candidates(semester,year,exam_type)
    if not candidates:
        candidates=static_collect_notices(session,semester,year,exam_type)
    if not candidates:
        raise RuntimeError("The DIU Notice Board could not be read. Please try again in a moment.")

    routines=_inspect_routine_documents(session,candidates,section,exam_type,semester,year)
    if not routines:
        raise RuntimeError(
            f"The DIU Notice Board was reached, but no {semester.title()} {year or ''} CSE {exam_type.title()} "
            f"routine containing batch {normalize_batch(section)} was verified. "
            f"Try again shortly, or check that the official notice has been published."
        )

    # Prefer: more exams for this batch, "updated" notices, then higher score.
    batch = normalize_batch(section)

    def _routine_rank(x):
        title = norm(f"{x.get('title','')} {x.get('detail_title','')} {x.get('final_url','')} {x.get('url','')}")
        exams = x.get("parsed_routine", {}).get("exams", [])
        batch_exams = sum(1 for e in exams if normalize_batch(e.get("batch")) == batch)
        updated = 1 if "updated" in title.lower() else 0
        return (batch_exams, updated, x.get("score", 0))

    routine = max(routines, key=_routine_rank)

    # Seat plan is optional — never fail the whole request if it is missing or slow.
    seat = None
    if include_seat_plan:
        try:
            seat = _inspect_seat_plan(session, candidates, routine, section, exam_type, semester, year)
        except Exception:
            seat = None

    return {
        "routine": {
            "url": routine.get("final_url") or routine.get("url"),
            "title": routine.get("detail_title") or routine.get("title") or "CSE Examination Routine",
            "bytes": routine["bytes"],
            "file_type": routine["file_type"],
            "exam_type": exam_type,
            "semester": semester.title(),
            "year": year,
        },
        "seat_plan": ({
            "url": seat.get("final_url") or seat.get("url"),
            "title": seat.get("detail_title") or seat.get("title") or "CSE Examination Seat Plan",
            "bytes": seat["bytes"],
            "file_type": seat["file_type"],
        } if seat else None),
        "source": NOTICEBOARD_URL,
    }

