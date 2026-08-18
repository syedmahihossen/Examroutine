import re
from datetime import datetime
from collections import defaultdict
from typing import Any, Optional
import pdfplumber
from openpyxl import load_workbook


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def extract_session(text: str) -> dict:
    s = norm(text).replace("–", "-").replace("—", "-")
    m = re.search(r"\b(Spring|Summer|Fall|Autumn|Winter)\s*[- ]?\s*(20\d{2})\b", s, re.I)
    if not m:
        # Some notices use Spring-26 / Summer 26.
        m = re.search(r"\b(Spring|Summer|Fall)\s*[- ]?\s*(\d{2})\b", s, re.I)
        if m:
            return {"semester": m.group(1).title(), "year": 2000 + int(m.group(2))}
        return {"semester": "", "year": None}
    return {"semester": m.group(1).title(), "year": int(m.group(2))}


def clean_course_code(value: str) -> str:
    m = re.search(r"\b([A-Z]{2,8}\d{3,4})\b", str(value or "").upper())
    return m.group(1) if m else ""


def normalize_batch(value: str) -> str:
    s = norm(value).upper()
    m = re.match(r"^(\d{2,3})(?:[_-]|\b)", s)
    if m:
        return m.group(1)
    m = re.search(r"BATCH\s*[-_:]?\s*(\d{2,3})\b", s)
    return m.group(1) if m else ""


def normalize_section(value: str) -> str:
    s = norm(value).upper().replace(" ", "").replace("-", "_")
    if s.startswith("BATCH"):
        s = s[5:]
    return s


def parse_date(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = norm(value).replace("–", "-").replace("—", "-")
    patterns = (
        r"\b(\d{1,2})[/-](\d{1,2})[/-](20\d{2})\b",
        r"\b(20\d{2})[/-](\d{1,2})[/-](\d{1,2})\b",
        r"\b(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s*,?\s*(20\d{2})\b",
    )
    for pattern in patterns:
        m = re.search(pattern, s, re.I)
        if not m:
            continue
        try:
            if m.group(1).isdigit() and len(m.group(1)) == 4:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            elif m.group(2).isalpha():
                return datetime.strptime(m.group(0).replace(",", ""), "%d %b %Y").strftime("%Y-%m-%d")
            else:
                d, mo, y = map(int, m.groups())
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        except (ValueError, TypeError):
            pass
    return None


def weekday(date_iso: Optional[str]) -> str:
    if not date_iso:
        return ""
    return datetime.strptime(date_iso, "%Y-%m-%d").strftime("%A")


DEFAULT_SLOT_TIMES = {
    "A": "09:00 AM - 11:00 AM",
    "B": "12:00 PM - 02:00 PM",
    "C": "03:00 PM - 05:00 PM",
}


def slot_time(slot: str, slot_times: Optional[dict] = None) -> str:
    return (slot_times or {}).get((slot or "").upper()) or DEFAULT_SLOT_TIMES.get((slot or "").upper(), "")


def extract_slot_times(text: str) -> dict:
    s = norm(text).replace("–", "-").replace("—", "-")
    out = {}
    pattern = re.compile(
        r"\bSlot\s*([ABC])\s*[:\-]?\s*(\d{1,2}:\d{2}\s*(?:am|pm))\s*-\s*(\d{1,2}:\d{2}\s*(?:am|pm))",
        re.I,
    )
    for m in pattern.finditer(s):
        def fmt(v):
            return v.upper().replace(" ", "") if re.match(r"\d{1,2}:\d{2}\s*[AP]M", v, re.I) else v
        out[m.group(1).upper()] = f"{fmt(m.group(2))} - {fmt(m.group(3))}"
    return out


def page_date_slot(page):
    text = page.extract_text() or ""
    date = parse_date(text)
    m = re.search(r"\bSlot\s*:?[ \t]*([ABC])\b", text, re.I)
    return date, (m.group(1).upper() if m else ""), extract_slot_times(text)


# ===================== ROUTINE =====================

def parse_course_cell(cell: str):
    text = norm(cell)
    if not text:
        return None
    text = re.sub(r"\[\s*\d+\s*\]", "", text)
    m = re.match(r"^([A-Z]{2,8}\d{3,4})\s*:?(.*)$", text, re.I)
    if not m:
        return None
    code = m.group(1).upper()
    name = norm(m.group(2)).lstrip(": ")
    return (code, name or code)


def _table_header_row(row):
    text = " ".join(norm(x or "") for x in row)
    return bool(re.search(r"Slot\s*A", text, re.I)) and bool(re.search(r"Slot\s*B", text, re.I))


def _parse_table_exam_rows(table, page_no, pending, slot_times=None):
    # DIU CSE final routine layout: Date | Slot A Course | Batch | Slot B Course | Batch | Slot C Course | Batch
    slot_columns = [(1, 2, "A"), (3, 4, "B"), (5, 6, "C")]
    found = []
    current_date = None
    for raw in table:
        row = list(raw) + [None] * max(0, 7 - len(raw))
        row = row[:7]
        if _table_header_row(row):
            current_date = None
            continue
        d = parse_date(row[0])
        if d:
            current_date = d
        for course_idx, batch_idx, slot in slot_columns:
            course = parse_course_cell(row[course_idx] or "")
            batch = normalize_batch(row[batch_idx] or "")
            if not course or not batch:
                continue
            codes = [m.group(1).upper() for m in re.finditer(r"\b([A-Z]{2,8}\d{3,4})\b", row[course_idx] or "", re.I)]
            item = {
                "date": current_date,
                "day": weekday(current_date),
                "slot": slot,
                "time": slot_time(slot, slot_times),
                "course_code": course[0],
                "course_codes": list(dict.fromkeys(codes or [course[0]])),
                "course_name": course[1],
                "batch": batch,
                "source_page": page_no,
            }
            if current_date:
                found.append(item)
            else:
                pending.append(item)
    return found


def _parse_xlsx_exam_routine(path: str) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    exams = []
    metadata = {"semester": "", "year": None, "slot_times": {}}
    course_re = re.compile(r"\b([A-Z]{2,8}\d{3,4})\b", re.I)
    batch_re = re.compile(r"(?:batch\s*[-_:]?\s*|\()?(\d{2,3})(?:\))?", re.I)
    slot_header_re = re.compile(r"\bslot\s*([ABC])\b", re.I)

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        sheet_text = "\n".join(norm(v) for row in rows[:80] for v in row if v is not None)
        if not metadata["semester"]:
            metadata.update({k: v for k, v in extract_session(sheet_text).items() if v})
        metadata["slot_times"].update(extract_slot_times(sheet_text))

        max_cols = max((len(r) for r in rows), default=0)
        # Published DIU CSE XLSX routines normally use 10 columns:
        # Date | A Course | A Student Count | A Batch | B Course | B Student Count | B Batch | C Course | C Student Count | C Batch.
        # Older/simpler sheets use 7 columns: Date + 3 Course/Batch pairs.
        slot_cols = {"A": 1, "B": 4, "C": 7} if max_cols >= 10 else {"A": 1, "B": 3, "C": 5}
        current_date = None
        for row_no, raw in enumerate(rows, 1):
            vals = [norm(v) for v in raw]
            if not any(vals):
                continue
            row_text = " ".join(vals)
            d = parse_date(next((v for v in raw if isinstance(v, datetime)), "")) or parse_date(row_text)
            if d:
                current_date = d

            row_items = []
            for slot, course_col in slot_cols.items():
                if course_col >= len(vals):
                    continue
                cell = vals[course_col]
                m = course_re.search(cell)
                if not m:
                    continue
                code = m.group(1).upper()
                # Search the whole slot's nearby cells for Batch-NN, but avoid taking student count as batch.
                nearby = vals[course_col:min(len(vals), course_col + 4)]
                batch = ""
                # Prefer an explicit Batch-NN label.
                for candidate in nearby:
                    bm = re.search(r"batch\s*[-_:]?\s*(\d{2,3})", candidate, re.I)
                    if bm:
                        batch = bm.group(1)
                        break
                # In the 10-column DIU layout the batch is exactly two cells
                # after the course: Course | student-count | (Batch-NN).
                if not batch and course_col + 2 < len(vals):
                    candidate = vals[course_col + 2]
                    bm = re.fullmatch(r"\(?\s*(\d{2,3})\s*\)?", candidate)
                    if bm:
                        batch = bm.group(1)
                # In the compact 7-column layout batch is the next cell.
                if not batch and course_col + 1 < len(vals):
                    candidate = vals[course_col + 1]
                    bm = re.fullmatch(r"\(?\s*(\d{2,3})\s*\)?", candidate)
                    if bm:
                        batch = bm.group(1)
                if not batch:
                    continue
                title = norm(cell[m.end():]).lstrip(": -–—") or re.sub(r"\b[A-Z]{2,8}\d{3,4}\b", "", cell, count=1, flags=re.I).strip(" :-–—")
                row_items.append({
                    "date": current_date,
                    "day": weekday(current_date),
                    "slot": slot,
                    "time": slot_time(slot, metadata["slot_times"]),
                    "course_code": code,
                    "course_codes": [code],
                    "course_name": title or code,
                    "batch": batch,
                    "source_page": f"{ws.title}:{row_no}",
                })
            exams.extend(row_items)

    unique = {}
    for e in exams:
        if e.get("date") and e.get("batch"):
            unique[(e["date"], e["slot"], e["course_code"], e["batch"])] = e
    return {
        "semester": metadata["semester"], "year": metadata["year"],
        "slot_times": metadata["slot_times"],
        "exams": sorted(unique.values(), key=lambda x: (x["date"], x["slot"], x["course_code"])),
    }


def parse_exam_routine(path: str) -> dict:
    if path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return _parse_xlsx_exam_routine(path)
    exams, pending = [], []
    metadata = {"semester": "", "year": None, "slot_times": {}}
    with pdfplumber.open(path) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            session = extract_session(text)
            if not metadata["semester"] and session["semester"]:
                metadata.update(session)
            metadata["slot_times"].update(extract_slot_times(text))
            page_dates = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]20\d{2}\b", text)
            next_page_date = parse_date(page_dates[0]) if page_dates else None
            if pending and next_page_date:
                for item in pending:
                    item["date"] = next_page_date
                    item["day"] = weekday(next_page_date)
                    exams.append(item)
                pending.clear()
            for table in page.extract_tables():
                exams.extend(_parse_table_exam_rows(table, page_no, pending, metadata["slot_times"]))
    unique = {(e["date"], e["slot"], e["course_code"], e["batch"]): e for e in exams if e.get("date")}
    return {"semester": metadata["semester"], "year": metadata["year"], "slot_times": metadata["slot_times"], "exams": sorted(unique.values(), key=lambda x:(x["date"],x["slot"],x["course_code"]))}


# ===================== SEAT PLAN =====================

def parse_seat_plan(path: str) -> list[dict]:
    """Parse the DIU date-wise seat-plan PDF.

    The official layout is a table with:
    Dept | ID | Course | Tech. Int. | Section | Room No | Seat(s) | Total

    Course code is inherited across page breaks because DIU often starts a
    continuation page with a section row that omits the course columns.
    """
    if not path.lower().endswith(".pdf"):
        raise ValueError("Seat plan must be a PDF")

    allocations = []
    current = None
    last_date = None
    last_slot = ""

    with pdfplumber.open(path) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            page_date, page_slot, _slot_times = page_date_slot(page)
            page_date = page_date or last_date
            page_slot = page_slot or last_slot
            last_date, last_slot = page_date, page_slot

            for table in page.extract_tables():
                for raw in table:
                    row = list(raw) + [None] * max(0, 8 - len(raw))
                    row = row[:8]

                    course_code = clean_course_code(row[1] or "")
                    section_cell = norm(row[4] or "")
                    room_cell = norm(row[5] or "")
                    seats_cell = norm(row[6] or "")
                    total_cell = norm(row[7] or "")

                    if section_cell and re.match(r"^\d{2,4}[_-][A-Z0-9]+$", section_cell, re.I):
                        inherited_course = ""
                        if not course_code and current:
                            if current.get("date") == page_date and current.get("slot") == page_slot:
                                inherited_course = current.get("course_code", "")

                        current = {
                            "date": page_date,
                            "slot": page_slot,
                            "section": normalize_section(section_cell),
                            "course_code": course_code or inherited_course,
                            "rooms": [],
                            "total": int(total_cell) if total_cell.isdigit() else None,
                            "source_page": page_no,
                        }
                        allocations.append(current)

                    if current and room_cell and seats_cell.isdigit():
                        current["rooms"].append({
                            "room": room_cell,
                            "seats": int(seats_cell),
                        })

    for a in allocations:
        seen = set()
        rooms = []
        for r in a["rooms"]:
            key = (r["room"], r["seats"])
            if key not in seen:
                seen.add(key)
                rooms.append(r)
        a["rooms"] = rooms
        if a["total"] is None:
            a["total"] = sum(r["seats"] for r in rooms)

    return allocations


# ===================== MATCHING =====================

def _dedupe_allocations(items):
    seen = set(); out = []
    for item in items:
        key = (item.get("source_page"), item.get("section"), item.get("date"), item.get("slot"), item.get("course_code"))
        if key not in seen:
            seen.add(key); out.append(item)
    return out


def _name_for_course_code(course_name: str, code: str, all_codes: list[str]) -> str:
    """Pick the title for one course when a routine cell lists several electives."""
    text = norm(course_name)
    code = clean_course_code(code)
    if not text or not code:
        return code or text

    # Split on common DIU separators: " / ", "/", " | "
    parts = re.split(r"\s*/\s*|\s*\|\s*", text)
    parts = [p for p in parts if p]
    if len(parts) <= 1 and len(all_codes or []) <= 1:
        # Single course cell: strip a leading "CODE:" if present.
        return re.sub(rf"^{re.escape(code)}\s*[:\-–—]?\s*", "", text, flags=re.I).strip() or text

    for part in parts:
        part_code = clean_course_code(part)
        if part_code == code:
            # "CSE441:UI and UX Design" or "Machine Learning" (first item may omit code)
            name = re.sub(rf"^{re.escape(code)}\s*[:\-–—]?\s*", "", part, flags=re.I).strip()
            return name or code

    # First segment often is "Machine Learning" for the first code without a prefix.
    if all_codes and clean_course_code(all_codes[0]) == code and parts:
        first = parts[0]
        if not clean_course_code(first):
            return first
        return re.sub(rf"^{re.escape(code)}\s*[:\-–—]?\s*", "", first, flags=re.I).strip() or code

    return code


def build_student_routine(routine: dict, seat_plan: Optional[list[dict]], section: str) -> dict:
    """Build a per-section routine.

    DIU often puts several elective course codes in one routine cell for a batch.
    Different sections of that batch take different electives. When a seat plan
    is available, the seat-plan course code is the source of truth for the
    section and we rewrite the displayed course to match it.
    """
    wanted_section = normalize_section(section)
    wanted_batch = normalize_batch(section)
    routine_exams = [e for e in routine.get("exams", []) if normalize_batch(e.get("batch")) == wanted_batch]
    allocations = [s for s in (seat_plan or []) if normalize_section(s.get("section")) == wanted_section]

    by_strong = defaultdict(list)
    by_date_course = defaultdict(list)
    by_date_slot = defaultdict(list)
    for s in allocations:
        code = clean_course_code(s.get("course_code", ""))
        date, slot = s.get("date"), (s.get("slot") or "").upper()
        if code:
            by_strong[(date, slot, code)].append(s)
            by_date_course[(date, code)].append(s)
        by_date_slot[(date, slot)].append(s)

    result = []
    warnings = []
    for exam in routine_exams:
        codes = list(dict.fromkeys(
            clean_course_code(c) for c in (exam.get("course_codes") or [exam.get("course_code", "")]) if clean_course_code(c)
        ))
        if not codes and exam.get("course_code"):
            codes = [clean_course_code(exam["course_code"])]

        candidates = []
        method = ""
        matched_code = ""

        # Prefer the course code that actually appears in this section's seat plan.
        for code in codes:
            strong = _dedupe_allocations(by_strong.get((exam["date"], exam["slot"], code), []))
            if strong:
                candidates = strong
                method = "date+slot+course"
                matched_code = code
                break

        if not candidates:
            for code in codes:
                c = _dedupe_allocations(by_date_course.get((exam["date"], code), []))
                if len(c) == 1:
                    candidates = c
                    method = "date+course"
                    matched_code = code
                    break

        if not candidates:
            c = _dedupe_allocations(by_date_slot.get((exam["date"], exam["slot"]), []))
            # Unique seat row for this section on that date/slot — use its course.
            if len(c) == 1:
                candidates = c
                method = "date+slot"
                matched_code = clean_course_code(c[0].get("course_code", "")) or (codes[0] if codes else "")

        chosen = candidates[0] if len(candidates) == 1 else None

        # Section-specific course: seat plan wins when it provides a code.
        display_code = matched_code or (codes[0] if codes else clean_course_code(exam.get("course_code", "")))
        if chosen and clean_course_code(chosen.get("course_code", "")):
            display_code = clean_course_code(chosen["course_code"])

        display_name = _name_for_course_code(
            exam.get("course_name", ""),
            display_code,
            codes or [display_code],
        )

        item = {
            **exam,
            "section": wanted_section,
            "course_code": display_code,
            "course_codes": [display_code] if display_code else codes,
            "course_name": display_name,
            "rooms": chosen["rooms"] if chosen else [],
            "total_students": chosen["total"] if chosen else None,
            "seat_match": bool(chosen),
            "seat_match_method": method,
            "seat_source_page": chosen.get("source_page") if chosen else None,
        }
        result.append(item)
        if seat_plan and not chosen:
            warnings.append(f"{exam['date']} {display_code or exam.get('course_code')}: seat allocation not matched.")

    # If seat plan lists a section exam that the batch routine cell omitted, still
    # surface nothing extra — routine remains the schedule authority for dates/slots.

    if not routine_exams:
        warnings.append(f"No routine entries found for batch {wanted_batch}.")
    if seat_plan and not allocations:
        warnings.append(f"Section {wanted_section} was not found in the seat-plan PDF.")

    for x in result:
        if x["seat_match"]:
            calc = sum(int(r.get("seats", 0)) for r in x["rooms"])
            total = x.get("total_students")
            if total is not None and calc != int(total):
                warnings.append(
                    f"{x['date']} {x['course_code']}: room seats total {calc}, PDF total says {total}."
                )

    return {
        "semester": routine.get("semester", ""),
        "year": routine.get("year"),
        "section": wanted_section,
        "batch": wanted_batch,
        "exam_count": len(result),
        "seat_plan_uploaded": bool(seat_plan),
        "matched_seat_count": sum(1 for x in result if x["seat_match"]),
        "exams": result,
        "warnings": warnings,
    }
